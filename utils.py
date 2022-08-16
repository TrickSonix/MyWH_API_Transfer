import os
import re
import datetime as dt
from setup import COMPANY_TYPES, DEFAULT_EXPENSE_ITEM, NDS_COMPARE, DEFAULT_GROUP, DEFAULT_TAG, DEFAULT_UOM, DEFAULT_CURRENCY, DEFAULT_RETAIL_CUSTOMER, DEFAULT_ORGANIZATION_ACCOUNTS, DEFAULT_ORGANIZATION_RETAIL_ACC, DEFAULT_STATE, DEFAULT_STORE, DEFAULT_SERVICE
from openpyxl import load_workbook, Workbook

def damerau_levenshtein_distance(s1, s2):
    d = {}
    lenstr1 = len(s1)
    lenstr2 = len(s2)
    for i in range(-1,lenstr1+1):
        d[(i,-1)] = i+1
    for j in range(-1,lenstr2+1):
        d[(-1,j)] = j+1
 
    for i in range(lenstr1):
        for j in range(lenstr2):
            if s1[i] == s2[j]:
                cost = 0
            else:
                cost = 1
            d[(i,j)] = min(
                           d[(i-1,j)] + 1, # deletion
                           d[(i,j-1)] + 1, # insertion
                           d[(i-1,j-1)] + cost, # substitution
                          )
            if i and j and s1[i]==s2[j-1] and s1[i-1] == s2[j]:
                d[(i,j)] = min (d[(i,j)], d[i-2,j-2] + 1) # transposition
 
    return d[lenstr1-1,lenstr2-1]

def find_max_look_like(string, string_list):
    res = []
    for s in string_list:
        alike = damerau_levenshtein_distance(string, s)
        res.append((s, alike))
    return sorted(res, key=lambda x: x[1])

def is_guid(s: str) -> bool:
    if isinstance(s, str) and re.match(r'^[a-z0-9]{8}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{12}$', s) and s != '00000000-0000-0000-0000-000000000000':
        return True
    else:
        return False

def transform_logs_into_excel(file_path, db):
        wb = Workbook()
        ws = wb.active
        ws.cell(1,1, value="Document type")
        ws.cell(1, 2, value="Document number")
        ws.cell(1,3, value='Document Date')
        ws.cell(1,4, value='Problem')
        row_count = 2
        with open(file_path, 'r', encoding='utf-8')as f:
            for n, line in enumerate(f.readlines()):
                splitted_line = line.split(' ')
                if 'ERROR' in line:
                    db_item = next(db.find_by_guid(next(filter(is_guid, splitted_line))))
                    ws.cell(row_count, 1, value=db_item['#type'].split('.')[-1])
                    ws.cell(row_count, 2, value=db_item['#value']['Number'])
                    ws.cell(row_count, 3, value=db_item['#value']['Date'])
                if n%2==1:
                    ws.cell(row_count, 4, value=line)
                    row_count +=1
        
        file_name = f'Excel_{file_path.split("/")[-1].split(".")[0]}.xlsx'
        wb.save(filename=file_name)

def make_error_report_from_log(log_path, time_mark, db=None):
    report_lines = []
    write_mark = False
    cur_time = time_mark
    with open(log_path, 'r', encoding='utf-8') as f:
        for line in f.readlines():
            line_splitted = line.split(' ')
            try:
                cur_time = dt.datetime.strptime(' '.join(line_splitted[:2])[:-4], "%Y-%m-%d %H:%M:%S")
                write_mark = False
            except ValueError:
                if write_mark:
                    report_lines.append(line)
            if cur_time >= time_mark:
                if 'ERROR' in line:
                    report_lines.append(line)
                    write_mark = True
    file_name = f'error_report_{log_path.split("/")[-1].split(".")[0]}_from_{dt.datetime.now().strftime("%Y-%m-%d-%H-%M")}.log'
    new_dir = dt.datetime.now().strftime("%Y-%m-%d")
    if not os.path.exists(f'logs/error_reports_{new_dir}'):
        os.mkdir(f'logs/error_reports_{new_dir}')
    with open(f'logs/error_reports_{new_dir}/{file_name}', 'w', encoding='utf-8') as f:
        for line in report_lines:
            f.write(line)
    a = input(f'Transform {file_name} into excel?\n')
    if a == 'y' and db:
        transform_logs_into_excel(f'logs/error_reports_{new_dir}/{file_name}', db)
    elif a=='y' and not db:
        print('DB GDE?')

def make_refs_list_from_error_log(error_log_path, in_file=False):
    refs_list = []
    with open(error_log_path, 'r', encoding='utf-8') as f:
        for line in f.readlines():
            refs_list.append(next(filter(is_guid, line.split(' '))))
    if in_file:
        file_name = error_log_path.split('/')[-1].split('.')[0]
        with open(f'{file_name}_refs.txt', 'w', encoding='utf-8'):
            for line in refs_list:
                f.write(line)
    else:
        return refs_list



class JSONPermute:

    def __init__(self, item: dict, item_type: str, db) -> None:
        self.item = item
        self.item_type = item_type
        self.db = db

    def date_convert(self, date):
        return dt.datetime.strptime(date, "%Y-%m-%dT%H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")

    def get_meta_by_guid(self, guid):
        item = next(self.db.find_by_guid(guid))
        return item.get('#mywh', {}).get('meta', {})

    def _find_date_price_by_guid(self, guid, date, price_type):
        item_price = 100
        max_date = None
        if not is_guid(price_type):
            price_type = "a0d058b3-323c-11e8-80d9-00505691ab3c"
        items = self.db.get_items("Документы", {"#type": "jcfg:DocumentObject.УстановкаЦенНоменклатуры", "#value.Товары": {"$elemMatch": {"Номенклатура": guid, "ВидЦены": price_type, "Цена": {"$gt": 0}}}})
        for doc in items:
            if not doc:
                return item_price
            doc_time = doc['#value'].get('Date', dt.datetime.now())
            if dt.datetime.strptime(doc_time, "%Y-%m-%dT%H:%M:%S") > dt.datetime.strptime(date, "%Y-%m-%dT%H:%M:%S"):
                continue
            else:
                if not max_date:
                    max_date = doc_time
                    item_price = next(filter(lambda x: x.get('Номенклатура') == guid and x.get('ВидЦены') == price_type, doc['#value']['Товары']))['Цена']
                else:
                    if doc_time > max_date:
                        max_date = doc_time
                        item_price = next(filter(lambda x: x.get('Номенклатура') == guid and x.get('ВидЦены') == price_type, doc['#value']['Товары']))['Цена']
        return item_price

    def _permute_organization(self):
        result = {}
        result['name'] = self.item['#value'].get('Description') or ""
        result['companyType'] = 'entrepreneur'
        result['inn'] = self.item['#value'].get('ИНН') or ""
        result['okpo'] = self.item['#value'].get('КодПоОКПО') or ""
        result['ogrnip'] = self.item['#value'].get('ОГРН') or ""
        result['group'] = DEFAULT_GROUP
        result['accounts'] = []
        accounts = self.db.get_items('Справочники', {"#type": "jcfg:CatalogObject.БанковскиеСчетаОрганизаций", "#value.Owner.#value": self.item['#value']['Ref']})
        for acc in accounts:
            temp = {}
            if acc.get('#mywh'):
                if acc['#mywh'].get('meta'):
                    temp['meta'] = acc['#mywh']['meta']
            temp['accountNumber'] = acc['#value']['НомерСчета']
            if not acc.get('#mywh'):
                temp['bankLocation'] = acc['#value']['Ref']
            bankName = (next(self.db.find_by_guid(acc['#value'].get('Банк'))).get('#value') or dict())
            if bankName:
                temp['bankName'] = bankName.get('Description') or ""
            result['accounts'].append(temp)
            del temp
        cash_acc = self.db.get_items('Справочники', {"#type": "jcfg:CatalogObject.Кассы", "#value.Owner.#value": self.item['#value']['Ref']})
        for acc in cash_acc:
            if not acc:
                continue
            temp = {}
            if acc.get('#mywh'):
                if acc['#mywh'].get('meta'):
                    temp['meta'] = acc['#mywh']['meta']
            temp['accountNumber'] = acc['#value']['Description']
            if not acc.get('#mywh'):
                temp['bankLocation'] = acc['#value']['Ref']
            result['accounts'].append(temp)
            del temp
        return result
    
    def _permute_product(self):
        result = {}
        result['name'] = self.item['#value'].get('Description', "")
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['uom'] = DEFAULT_UOM
        result['group'] = DEFAULT_GROUP
        result['article'] = self.item['#value'].get('Артикул') or ""
        result['shared'] = True
        # result['salePrice'] = []
        # for price in self.db.get_items('Справочники', {'#type': "jcfg:CatalogObject.ВидыЦен", "#value.IsFolder": False}):
        #     temp = {}
        #     temp['value'] = self._find_date_price_by_guid(guid=self.item["#value"]["Ref"], date=dt.datetime.now(), price_type=price['#value']['Ref'])
        #     if price.get('#mywh', {}).get('meta', {}):
        #         temp['priceType']  = {'meta': price.get('#mywh', {}).get('meta', {})}
        #     else:
        #         temp['priceType'] = price.get('#value', {}).get('Description', '')
        #     temp['currency'] = DEFAULT_CURRENCY
        #     result.append(temp)
        #     del temp
        return result
    
    def _permute_counterparty(self):
        result = {}
        result['name'] = self.item['#value'].get('Description') or ""
        result['companyType'] = COMPANY_TYPES.get(self.item['#value'].get('ЮрФизЛицо'))
        result['inn'] = self.item['#value'].get('ИНН') or ""
        result['okpo'] = self.item['#value'].get('КодПоОКПО') or ""
        result['kpp'] = self.item['#value'].get('КПП') or ""
        result['shared'] = True
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['tag'] = [DEFAULT_TAG, 'Выгружено из 1С']
        result['group'] = DEFAULT_GROUP
        result['accounts'] = [{'accountNumber': 'Счет для документов без счета', 'isDefault': True}]
        accounts = self.db.get_items('Справочники', {"#type": "jcfg:CatalogObject.БанковскиеСчетаКонтрагентов", "#value.Owner.#value": self.item['#value']['Ref']})
        # for contact in self.item['#value'].get('КонтактнаяИнформация', []): #потом доделаю адреса
        #     if contact.get('Тип') == 'Адрес':

        for acc in accounts:
            if not acc:
                continue
            temp = {}
            if acc.get('#mywh'):
                if acc['#mywh'].get('meta'):
                    temp['meta'] = acc['#mywh']['meta']
            temp['accountNumber'] = acc['#value']['НомерСчета']
            if not acc.get('#mywh'):
                temp['bankLocation'] = acc['#value']['Ref']
            bankName = (next(self.db.find_by_guid(acc['#value'].get('Банк'))).get('#value') or dict())
            if bankName:
                temp['bankName'] = bankName.get('Description') or ""
            result['accounts'].append(temp)
            del temp
        return result

    def _permute_expenseitem(self):
        result = {}
        result['name'] = '-'.join([self.item['#value'].get('Description') or self.item['#value']['Ref'], "CG"])
        result['description'] = self.item['#value'].get('Описание') or ""
        result['group'] = DEFAULT_GROUP
        return result

    def _permute_store(self):
        result = {}
        result['name'] = self.item['#value'].get('Description') or ""
        result['group'] = DEFAULT_GROUP
        return result

    def _permute_supply(self):
        result = {}
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value']['Контрагент'])}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value']['Организация'])}
        organization_account_meta = self.get_meta_by_guid(self.item['#value']['БанковскийСчетОрганизации'])
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value']['Склад'])}
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = pos['Количество']
            temp['vat'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[0]
            temp['vatEnabled'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[1]
            result['positions'].append(temp)
        del temp
        return result

    def _permute_enter(self):
        result = {}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = "\n".join([self.item['#value'].get('Комментарий'), f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}']).strip()
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = pos['Количество']
            result['positions'].append(temp)
        del temp
        return result

    def _permute_purchasereturn(self):
        result = {}
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетОрганизации'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))} 
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = pos['Количество']
            temp['vat'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[0]
            temp['vatEnabled'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[1]
            result['positions'].append(temp)
            del temp
        supply = self.get_meta_by_guid(self.item['#value'].get('ДокументПоступления'))
        if supply:
            result['supply'] = {'meta': supply} 
        return result

    def _permute_customerorder(self):
        result = {}
        result['name'] = ''.join(['CG-1C-', self.item['#value']['Number'].strip()])
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетОрганизации'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['state'] = DEFAULT_STATE
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = pos['Количество']
            temp['vat'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[0]
            temp['vatEnabled'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[1]
            temp['reserve'] = pos['Количество'] if pos['ВариантОбеспечения'] == 'Отгрузить' else 0
            result['positions'].append(temp)
            del temp
        return result

    def _permute_demand(self):
        result = {}
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетОрганизации'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = pos['Количество']
            temp['vat'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[0]
            temp['vatEnabled'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[1]
            result['positions'].append(temp)
            del temp
        customerOrder = self.get_meta_by_guid(self.item['#value'].get('ЗаказКлиента', {}).get('#value'))
        if customerOrder:
            result['customerOrder'] = {'meta': customerOrder}
        return result

    def _permute_salesreturn(self):
        result = {}
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетОрганизации'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = pos['Количество']
            temp['vat'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[0]
            temp['vatEnabled'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[1]
            result['positions'].append(temp)
            del temp
        demand = self.get_meta_by_guid((self.item.get('#value', {}).get('ДокументРеализации', {}) or {}).get('#value', {}))
        if demand:
            result['demand'] = {'meta': demand}
        return result

    def _permute_loss(self):
        result = {}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        expenseItem = self.get_meta_by_guid(self.item['#value']['СтатьяРасходов']).get('#value', {}).get('Description', '')
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(self._find_date_price_by_guid(pos.get('Номенклатура'), self.item.get('#value', {}).get('Date', ""), self.item['#value']['ВидЦены'])*60, 0))
            temp['quantity'] = pos['Количество']
            temp['reason'] = expenseItem
            result['positions'].append(temp)
            del temp
        return result

    def _permute_paymentin(self):
        result = {}
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетОрганизации'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta} 
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])    
        if self.item['#value']['ХозяйственнаяОперация'] == "ПоступлениеДенежныхСредствСДругогоСчета":
            result['agent'] = result['organization']
            agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетОтправитель'))
            if agent_account_meta:
                result['agentAccount'] = {'meta': agent_account_meta}
        else:
            result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
            agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента', {}))
            if agent_account_meta:
                result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['shared'] = True
        result['sum'] = int(round(self.item['#value'].get('СуммаДокумента', 0)*100, 0))
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['paymentPurpose'] = self.item['#value'].get('НазначениеПлатежа', '')
        result['operations'] = []
        for pos in self.item['#value'].get('РасшифровкаПлатежа', []):
            temp = {}
            payment_base_meta = self.get_meta_by_guid(pos.get('УдалитьЗаказ', {}) or {}).get('#value', {}) or self.get_meta_by_guid((pos.get('ОснованиеПлатежа', {}) or {}).get('#value', {}))
            if payment_base_meta:
                temp['meta'] = payment_base_meta
                result['operations'].append(temp)
            del temp
        if not result['operations']:
            result.pop('operations')
        return result

    def _permute_paymentout(self): 
        result = {}
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчет'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        if self.item['#value']['ХозяйственнаяОперация'] == "ПеречислениеДенежныхСредствНаДругойСчет":
            result['agent'] = result['organization']
            agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетПолучатель'))
            if agent_account_meta:
                result['agentAccount'] = {'meta': agent_account_meta}
        elif self.item['#value']['ХозяйственнаяОперация'] == "ВыплатаЗарплатыНаЛицевыеСчета":
            result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('ПодотчетноеЛицо'))}
            agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
            if agent_account_meta:
                result['agentAccount'] = {'meta': agent_account_meta}
        else:
            result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
            agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
            if agent_account_meta:
                result['agentAccount'] = {'meta': agent_account_meta}
        
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['group'] = DEFAULT_GROUP
        result['shared'] = True
        result['sum'] = int(round(self.item['#value'].get('СуммаДокумента', 0)*100, 0))
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['paymentPurpose'] = self.item['#value'].get('ФорматированноеНазначениеПлатежа', '')
        result['operations'] = []
        if self.item['#value']['ХозяйственнаяОперация'] == "ПеречислениеДенежныхСредствНаДругойСчет":
            result['expenseItem'] = {'meta': self.get_meta_by_guid("639c7631-7669-11e5-a965-3085a9eabb90")}
        for pos in self.item['#value'].get('РасшифровкаПлатежа', []):
            if not result.get('expenseItem'):
                ref = pos['СтатьяРасходов']['#value']
                if is_guid(ref):
                    result['expenseItem'] = {'meta': self.get_meta_by_guid(pos['СтатьяРасходов']['#value'])}
            new_string = ' '.join([pos.get('Содержание', ""), 'Сумма', str(max(pos['Сумма'], pos.get('СуммаСНДС', 0)))]).strip()
            result['paymentPurpose'] = '\n'.join([new_string, result['paymentPurpose']]).strip()
            if (pos.get('УдалитьЗаказ', {}) or {}).get('Ref', {}) != self.item["#value"]['Ref']:
                operation_meta = self.get_meta_by_guid(pos.get('УдалитьЗаказ', {}) or {}).get('#value', {}) or self.get_meta_by_guid((pos.get('ОснованиеПлатежа', {}) or {}).get('#value', {}))
                if operation_meta:
                    result['operations'].append({'meta': operation_meta})
        if not result.get('expenseItem'):
            result['expenseItem'] = DEFAULT_EXPENSE_ITEM
        if not result.get('operations'):
            result.pop('operations')
                    
        return result

    def _permute_move(self):
        result = {}
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        result['sourceStore'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('СкладОтправитель'))}
        result['targetStore'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('СкладПолучатель'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['quantity'] = pos['Количество']
            result['positions'].append(temp)
            del temp
        return result

    def _permute_retaildemand(self):
        result = {}
        result['agent'] = DEFAULT_RETAIL_CUSTOMER #здесь должна быть meta для контрагента "Розничный покупатель"
        agent_account_meta = {} 
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta} #мета счета "Розничный покупатель"
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))} 
        result['organizationAccount'] = DEFAULT_ORGANIZATION_RETAIL_ACC.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            if pos['Количество'] < 0:
                continue
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = pos['Количество']
            temp['vat'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[0]
            temp['vatEnabled'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[1]
            result['positions'].append(temp)
            del temp
        return result

    def _permute_salesreturn_from_retail(self):
        result = {}
        result['agent'] = DEFAULT_RETAIL_CUSTOMER
        agent_account_meta = {}
        result['group'] = DEFAULT_GROUP
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        result['organizationAccount'] = DEFAULT_ORGANIZATION_RETAIL_ACC.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            if pos['Количество'] > 0:
                continue
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/pos['Количество'], 0))
            temp['quantity'] = abs(pos['Количество'])
            temp['vat'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[0]
            temp['vatEnabled'] = (NDS_COMPARE.get(pos['СтавкаНДС'], [0, False]))[1]
            result['positions'].append(temp)
            del temp
        return result

    def _permute_supply_from_services_purchase(self):
        result = {}
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value']['Контрагент'])}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value']['Организация'])}
        organization_account_meta = self.get_meta_by_guid(self.item['#value']['БанковскийСчетОрганизации'])
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['store'] = DEFAULT_STORE
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        for pos in self.item['#value'].get('Расходы', []):
            temp = {}
            temp['assortment'] = DEFAULT_SERVICE
            temp['price'] = int(round(max(pos['Сумма'], pos.get('СуммаСНДС', 0))*100/(pos['Количество'] or 1), 0))
            temp['quantity'] = pos['Количество'] or 1
            result['positions'].append(temp)
            result['description'] = '\n'.join([pos['Содержание'] , result['description']]).strip()
        del temp
        return result

    def _permute_loss_from_internal(self):
        result = {}
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        result['shared'] = True
        result['store'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Склад'))}
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['positions'] = []
        for pos in self.item['#value'].get('Товары', []):
            temp = {}
            temp['assortment'] = {'meta': self.get_meta_by_guid(pos.get('Номенклатура'))}
            temp['price'] = int(round(self._find_date_price_by_guid(pos.get('Номенклатура'), self.item['#value'].get('Date', ""), self.item['#value']['ВидЦены'])*60, 0))
            temp['quantity'] = pos['Количество']
            temp['reason'] = next(self.db.find_by_guid(pos['СтатьяРасходов']['#value'])).get('#value', {}).get('Description', '')
            result['positions'].append(temp)
            del temp
        return result

    def _permute_paymentin_from_orders(self):
        result = {}
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta} #сюда возможно придется впилить какой-нибудь счет по-умолчанию
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('Касса'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['sum'] = int(round(self.item['#value'].get('СуммаДокумента', 0)*100, 0))
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['operations'] = []
        for pos in self.item['#value'].get('РасшифровкаПлатежа', []):
            temp = {}
            payment_base_meta = self.get_meta_by_guid(pos.get('УдалитьЗаказ', {}) or {}).get('#value', {}) or self.get_meta_by_guid((pos.get('ОснованиеПлатежа', {}) or {}).get('#value', {}))
            if payment_base_meta:
                temp['meta'] = payment_base_meta
                result['operations'].append(temp)
            del temp
        if not result['operations']:
            result.pop('operations')
        return result

    def _permute_paymentout_from_orders(self):
        result = {}
        result['agent'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Контрагент'))}
        agent_account_meta = self.get_meta_by_guid(self.item['#value'].get('БанковскийСчетКонтрагента'))
        if agent_account_meta:
            result['agentAccount'] = {'meta': agent_account_meta} #счет по умолчанию впили
        result['applicable'] = self.item['#value'].get('Posted', False)
        result['moment'] = self.date_convert(self.item['#value'].get('Date', ""))
        result['organization'] = {'meta': self.get_meta_by_guid(self.item['#value'].get('Организация'))}
        organization_account_meta = self.get_meta_by_guid(self.item['#value'].get('Касса'))
        if organization_account_meta:
            result['organizationAccount'] = {'meta': organization_account_meta}
        elif DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация']):
            result['organizationAccount'] = DEFAULT_ORGANIZATION_ACCOUNTS.get(self.item['#value']['Организация'])
        result['shared'] = True
        result['sum'] = 0
        result['paymentPurpose'] = ''
        result['description'] = f'Документ создан автоматически на основании {self.item["#value"]["Ref"]}'
        result['group'] = DEFAULT_GROUP
        result['operations'] = []
        for pos in self.item['#value'].get('РасшифровкаПлатежа', []):
            if not result.get('expenseItem'):
                if is_guid(pos['СтатьяРасходов']['#value']):
                    result['expenseItem'] = {'meta': self.get_meta_by_guid(pos['СтатьяРасходов']['#value'])}
                else:
                    result['expenseItem'] = {'meta': self.get_meta_by_guid('639c7631-7669-11e5-a965-3085a9eabb90')}
            new_string = ' '.join([pos['Комментарий'], 'Сумма', str(max(pos['Сумма'], pos.get('СуммаСНДС', 0)))])
            result['paymentPurpose'] = '\n'.join([new_string, result['paymentPurpose']]).strip()
            result['sum'] += int(round(pos.get('СуммаДокумента', 0)*100, 0))
            operation = self.get_meta_by_guid(pos.get('УдалитьЗаказ', {}) or {}).get('#value', {}) or self.get_meta_by_guid((pos.get('ОснованиеПлатежа', {}) or {}).get('#value', {}))
            if operation:
                result['operations'].append({'meta': operation})
        return result
    
    def permute(self, update_meta=False, **kwargs) -> dict:
        if not self.item:
            return {}
        result = {}
        if self.item_type == 'organization':
            result.update(self._permute_organization())
        if self.item_type == 'counterparty':
            result.update(self._permute_counterparty())
        if self.item_type == 'product':
            result.update(self._permute_product())
        if self.item_type == 'service':
            result.update(self._permute_product())
        if self.item_type == 'counterparty':
            result.update(self._permute_counterparty())
        # if self.item_type == 'counterpartyadjustment':
        #     result.update(self._permute_counterpartyadjustment())
        if self.item_type == 'expenseitem':
            result.update(self._permute_expenseitem())
        if self.item_type == 'store':
            result.update(self._permute_store())
        if self.item_type == 'supply':
            result.update(self._permute_supply())
        if self.item_type == 'enter':
            result.update(self._permute_enter())
        if self.item_type == 'move':
            result.update(self._permute_move())
        if self.item_type == 'purchasereturn':
            result.update(self._permute_purchasereturn())
        if self.item_type == 'customerorder':
            result.update(self._permute_customerorder())
        if self.item_type == 'demand':
            result.update(self._permute_demand())
        if self.item_type == 'salesreturn':
            result.update(self._permute_salesreturn())
        if self.item_type == 'loss':
            result.update(self._permute_loss())
        if self.item_type == 'paymentin':
            result.update(self._permute_paymentin())
        if self.item_type == 'paymentout':
            result.update(self._permute_paymentout())
        # if self.item_type == 'cashin':
        #    result.update(self._permute_cashin())
        # if self.item_type == 'cashout':
        #     result.update(self._permute_cashout())
        if self.item_type == 'retaildemand':
            result.update(self._permute_retaildemand())
        if self.item_type == 'supply_from_services_purchase':
            result.update(self._permute_supply_from_services_purchase())
        if self.item_type == 'salesreturn_from_retail':
            result.update(self._permute_salesreturn_from_retail())
        if self.item_type == 'loss_from_internal':
            result.update(self._permute_loss_from_internal())
        if self.item_type == 'paymentin_from_orders':
            result.update(self._permute_paymentin_from_orders())
        if self.item_type == 'paymentout_from_orders':
            result.update(self._permute_paymentout_from_orders())
        if self.item.get('#mywh', {}).get('meta', None) and not update_meta:
            result['meta'] = self.item['#mywh'].get('meta')
        if kwargs.get('update_items'):
            to_return = {}
            for key in result.keys():
                if key in kwargs.get('update_items') or key == '#mywh':
                    to_return[key] = result[key]
            if not self.item['#mywh'].get('meta'):
                return False
            to_return['meta'] = self.item['#mywh'].get('meta')
            return to_return

        return result



# def get_df_from_file(path):
#     sheet_names = pd.ExcelFile(path).sheet_names
#     res = {}
#     for sheet_name in sheet_names:
#         if sheet_name == 'Summary':
#             continue
#         for cols_seq in ['A,B,F,K,L', 'A,B,D,E,G:J', 'M,N,P:S', 'T,U,Y,Z']:
#             df = pd.read_excel(path, sheet_name=sheet_name, usecols=cols_seq, header=5)
#             df.columns = [re.sub(r'\.\d+', '', re.sub(r'Unnamed: \d+', '', ' '.join(x).strip())) for x in zip(df.iloc[0].fillna('').values, df.columns)]
#             df = df.dropna()
#             if len(df) > 0:
#                 if res.get(sheet_name):
#                     res[sheet_name].append(df.copy())
#                 else:
#                     res[sheet_name] = [df.copy()]
#             else:
#                 continue
#     return res

# def product_code(code):
#     if 'CGA' in code:
#         return 'CGA' + code.replace('CGA', '')
#     else:
#         return code

# def dfs_to_excel(path):
#     dfs = get_df_from_file(path)
#     path_splitted = path.split('/')
#     end_path = '/'.join(path_splitted[:-1] + [''.join(path_splitted[-1].split('.')[:-1])])
#     try:
#         os.mkdir(end_path)
#     except OSError as e:
#         print(e)
#     for key in dfs.keys():
#         for i, d in enumerate(dfs[key]):
#             d['ProdCode'] = d['ProdCode'].astype('str').apply(product_code)
#             d.to_excel(f'{end_path}/{key}_{i}.xlsx', index=False)

if __name__ == '__main__':
    print(find_max_look_like('Серп', ['Настольная игра Серп', 'Серп. Дополнение']))
    pass