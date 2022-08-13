from ftplib import error_perm
import requests
import os
import sys
import time
from openpyxl import load_workbook, Workbook
import datetime as dt
from setup import COLLECTIONS_COMPARE, REQUEST_LINKS, PSWD, USERNAME, HEADERS, CUSTOMERORDER_COPIED_FIELDS, MAIN_URL, MAX_DATA_SIZE, REQUESTS_SECONDS_RESTRICTION, REQUESTS_LIMIT, TIMEOUT_TIME, MAX_ITEM_COUNT, OWNER
import json
import functools
import operator
from logger import setup_logger
from logging import DEBUG, INFO
from ratelimit import limits
from ratelimit.exception import RateLimitException
from db_server import CrowdGamesDB
from utils import JSONPermute


"""Нужно будет протестировать изменения в коде и дальше уже пилить заполнение тела для разных запросов и распарса json'a из базы.
    Позже пойму где еще нужно добавить логгирование.
    Где-то еще можно добавить асинхронщины... Скорее всего в GET-запросах. Но если добавлять ее, то надо учитывать ограничения API моего склада (см документацию)
    
    Поскольку я таки пришел к тому что сделал БД для переноса, нужно будет туда напихать все справочники из моего склада, чтобы не делать запрос к апи каждый раз
    
    demand.json надо пофиксить, а скорее всего вообще все переделать."""




def sleep_and_retry(func):
    '''
    Return a wrapped function that rescues rate limit exceptions, sleeping the
    current thread until rate limit resets.
    :param function func: The function to decorate.
    :return: Decorated function.
    :rtype: function
    '''
    @functools.wraps(func)
    def wrapper(*args, **kargs):
        '''
        Call the rate limited function. If the function raises a rate limit
        exception sleep for the remaing time period and retry the function.
        :param args: non-keyword variable length argument list to the decorated function.
        :param kargs: keyworded variable length argument list to the decorated function.
        '''
        while True:
            try:
                return func(*args, **kargs)
            except RateLimitException as exception:
                basic_logger.info('Program stopped from requests restrictions')
                time.sleep(exception.period_remaining)
    return wrapper
    
class MyWHAPI:

    def __init__(self, username, pswd, db: CrowdGamesDB, headers, max_data_size, max_item_count, timeout_time) -> None:
        self.session = requests.Session()
        self.session.auth = (username, pswd)
        self.session.headers.update(headers)
        self.db = db
        self.max_data_size = max_data_size
        self.max_item_count = max_item_count
        self.timeout_time = timeout_time
            
    @staticmethod
    def list_flatten(list_2):
        return functools.reduce(operator.iconcat, list_2, [])
        
    @staticmethod
    def to_query(params):    
        """аргумент params передается в виде словаря, в котором все значения должны быть списками со значениями включающие оператор сравнения
        например: {'id=': 'lolkek'}"""
        return ';'.join([f'{key}{item}' for key, item in zip(MyWHAPI.list_flatten([[key]*len(value) for key, value in params.items()]), MyWHAPI.list_flatten([x for x in params.values()]))])
    
    def json_to_excel(self, json_data, **kwargs):
        #пока что явно нужно указывать колонки, но наверное хорошо бы сделать какое-то значение по-умолчанию
        data = self.json_load(json_data)
        wb = Workbook()
        ws = wb.active
        if kwargs.get('columns'):
            for n_row, row in enumerate(data):
                for n_column, column in enumerate(kwargs['columns']):
                    if n_row == 0:
                        ws.cell(row=n_row+1, column=n_column+1, value=column)
                    else:
                        if not isinstance(row.get(column), str):
                            continue
                        else:
                            ws.cell(row=n_row+1, column=n_column+1, value=row.get(column))
            filename = kwargs.get('save_path') or f'{dt.datetime.now().strftime("%d-%m-%y %H-%M")}.xlsx' 
            wb.save(filename)
            basic_logger.info(f'File {filename} was written.')


    @staticmethod
    def json_load(json_data):
        if json_data:
            if isinstance(json_data, dict) or isinstance(json_data, list):
                return json_data
            if os.path.exists(json_data):
                with open(json_data, 'r', encoding='utf-8') as f:
                    try:
                        return json.load(f)
                    except json.decoder.JSONDecodeError as e:
                        json_logger.error(e)
                        json_logger.debug(f'JSON_data: \n{json_data}')
                        return {}
            else:
                try:
                    return json.loads(json_data)
                except json.decoder.JSONDecodeError as e:
                    json_logger.error(e)
                    json_logger.debug(f'JSON_data: \n{json_data}')
                    return {}
        else:
            return {}

    @staticmethod
    def dump_json(data, file_path=None, **kwargs) -> str:
        try:
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, **kwargs)
                return 'Success'
            else:
                return json.dumps(data, **kwargs)
        except json.JSONDecodeError as e:
            json_logger.exception(f'{e}')
            json_logger.info(f'Decoding error, while trying to jsonize {data}')
            return ''

    def _check_data_size(self, data, appended_data, restriction):
        return sys.getsizeof(json.dumps(data+[appended_data])) >= restriction

    @sleep_and_retry
    @limits(calls=REQUESTS_LIMIT, period=REQUESTS_SECONDS_RESTRICTION)
    def request(self, method : str, path : str,  **kwargs) -> requests.Response:    
        try:
            with self.session as session:
                response = session.request(method=method, url=''.join([MAIN_URL, path or '']), **kwargs) #, timeout=self.timeout_time убрал по совету техпода моегосклада
                request_logger.info(f'URL: {response.url}. Method: {method}. Status code {response.status_code}')
                response.raise_for_status()
                return response
        except requests.exceptions.HTTPError as e:
            request_logger.debug(f'Not transfered or not fully transfered request body/')
            file_path=f'logs/json_body_err/{response.status_code}_body_{dt.datetime.now().strftime("%Y-%m-%d-%H:%M")}.json'
            self.dump_json(self.json_load(kwargs.get("data")), file_path=file_path, ensure_ascii=False, indent=4)
            request_logger.exception(f'{e}', exc_info=True)
            request_logger.debug(f'Response text: !{response.text}!')
            return response
        except requests.exceptions.Timeout as e:
            request_logger.debug(f'Timeout Error')
            request_logger.exception(f'{e}', exc_info=True)
            file_path=f'logs/json_body_err/{response.status_code}_body_{dt.datetime.now().strftime("%Y-%m-%d-%H:%M")}.json'
            self.dump_json(self.json_load(kwargs.get("data")), file_path=file_path, ensure_ascii=False, indent=4)
            return response
        except requests.exceptions.ConnectionError as e:
            request_logger.debug(f'Connection Error. Try to not CRY.')
            request_logger.exception(f'{e}', exc_info=True)
            file_path=f'logs/json_body_err/{response.status_code}_body_{dt.datetime.now().strftime("%Y-%m-%d-%H:%M")}.json'
            self.dump_json(self.json_load(kwargs.get("data")), file_path=file_path, ensure_ascii=False, indent=4)
            return requests.Response()
  
    def create_json_data_body(self, body_schema, schema_name, **kwargs):
        try:
            item_type = body_schema["#type"].split('.')[0].split(':')[-1]
            if kwargs.get('query'):
                new_query = kwargs.get('query')
                new_query.update({"#type": body_schema['#type']})
            else:
                new_query = {"#type": body_schema['#type']}
            if item_type == 'DocumentObject' and kwargs.get('only_posted_documents'):
                new_query.update({"#value.Posted": True})
            items = self.db.get_items(item_type, new_query)
        except Exception as e: #нужно будет позже понять какие ошибки тут отлавливать
            new_query = kwargs.get('query')
            basic_logger.exception(f'{e}', exc_info=True)
            basic_logger.debug(f'Body schema: {body_schema}. Query: {new_query}')
            items = []
        res = []
        for item in items:
            perm = JSONPermute(item, schema_name.lower(), self.db)
            permuted_data = perm.permute(update_meta=kwargs.get('update_meta', False))
            if not permuted_data:
                continue
            item_ref = item.get('#value', {}).get('Ref')
            if schema_name.lower() == 'organization' or schema_name.lower() == 'counterparty':
                max_items = 100
            else:
                max_items = self.max_item_count
            if self._check_data_size(res, (item_ref, permuted_data), self.max_data_size) or len(res)+1 > max_items:
                yield res
                res = []
            res.append((item_ref, permuted_data))
            ref_type = item['#type']
            json_logger.debug(f'{ref_type} item with ref {item_ref} was permuted into: \n{self.dump_json(permuted_data, ensure_ascii=False, indent=4)}')
        
        yield res
    
    def update_db_metas(self, zipped_data):
        error_count = 0
        for ref, resp in zipped_data:
            if resp.get('errors'):
                err = '\n'.join([x['error'] for x in resp['errors']])
                basic_logger.error(f'Error occured in response with {ref} item: \n{err}')
                error_count += 1
                continue
            meta = resp.get('meta')
            meta['href'] = meta['href'].split('?')[0]
            if meta:
                update = self.db.update_item(ref, meta)
                if update:
                    basic_logger.info(f'Success in update {ref} item')
                else:
                    basic_logger.info(f'Fail in update to DB {ref} item')
                    self.dump_json((ref, meta), file_path=f'logs/Connection_Error_body_{dt.datetime.now().strftime("%Y-%m-%d-%H-%M")}.json', ensure_ascii=False, indent=4)
                    db_logger.error(f'Pair ref and meta not written.')
                    error_count +=1
            else:
                basic_logger.info(f'Fail in update to DB {ref} item. Meta missed.')
                error_count +=1
            if resp.get('accounts'):
                for acc in resp['accounts'].get('rows'):
                    acc_meta = acc['meta']
                    acc_ref = acc.get('bankLocation')
                    if not acc_ref:
                        continue
                    update_acc = self.db.update_item(acc_ref, acc_meta)
                    if update_acc:
                        basic_logger.info(f'Success in update {acc_ref} item')
                    else:
                        basic_logger.info(f'Fail in update to DB {acc_ref} item')
                        db_logger.error(f'Next pair ref and meta not written: \n !{self.dump_json((acc_ref, acc_meta), ensure_ascii=False, indent=4)}!')
                        error_count +=1
        return error_count
    
    def restore_meta(self, data: list[dict], **kwargs):
        if not kwargs.get('path'):
            return [({}, {})]
        start_time = dt.datetime.now()
        href = kwargs.get('path')
        refs_list = [x[1].get('description', "missed").split(' ')[-1] for x in data]
        missed_meta = []
        error_count = 0
        while refs_list:
            for n, ref in enumerate(refs_list):
                entity = self.request('GET', path=href, params={'search': ref})
                if entity.status_code == 200:
                    entity_json = entity.json()
                    if entity_json['meta']['size'] == 1:
                        missed_meta.append((refs_list.pop(n), {'meta': entity_json['rows'][0]['meta']}))
                        basic_logger.info(f'Succesufully finded meta for {ref}')
                    elif entity_json['meta']['size'] == 0:
                        basic_logger.info(f'Did not finded any item with {refs_list.pop(n)}')
                    else:
                        if not kwargs.get('drop_duplicates'):
                            basic_logger.warning(f'For {refs_list.pop(n)} search matched more then 1 entity')
                        else:
                            basic_logger.info(f'Deleting duplicates')
                            self.ctrl_z(path=kwargs.get('path'), params={'limit': str(entity_json['meta']['size']-1), 'search': ref}, next_href=kwargs.get('next_href'))
                            
                else:
                    basic_logger.warning(f'In restore meta server got {entity.status_code}')
                    error_count += 1
                    if error_count >= 45:
                        break
        
            if dt.datetime.now() - start_time > dt.timedelta(minutes=90):
                basic_logger.warning(f'Failed to restore meta.')
                return [({}, {})]

        return missed_meta

    
    def migrate(self, schema, path, **kwargs):
        if schema.lower() == 'product':
            query = {"#value.ТипНоменклатуры": "Товар", "#value.IsFolder": False}
        elif schema.lower() == 'service':
            query = {"#value.ТипНоменклатуры": "Услуга", "#value.IsFolder": False}
        else:
            query = {}
        if kwargs.get('skip_existing'):
            query.update({"#mywh": {"$exists": False}})
        if kwargs.get('skip_deleted'):
            query.update({"#value.DeletionMark": False})
        body_schema = self.json_load(f'schemas/{schema.lower()}.json')
        data = self.create_json_data_body(body_schema=body_schema, query=query, schema_name=schema, **kwargs)
        error_count = 0
        for data_part in data:
            zipped_meta = None
            if schema.lower() == 'organization' or schema.lower() == 'counterparty':
                params = {'expand': 'accounts'}
            else:
                params = {}
            if not data_part:
                continue
            response = self.request(method='POST', path=path, data=self.dump_json([x[1] for x in data_part]), params=params)

            if response.status_code != 200 or response.status_code is None:
                basic_logger.warning(f'Something goes wrong with {schema} migration. Get {response.status_code} from server.')
                if response.status_code == 504 or response.status_code is None:
                    file_path= f'Data/Missed_meta_{dt.datetime.now().strftime("%Y-%m-%d-%H-%M")}.json'
                    self.dump_json(data_part, file_path=file_path, ensure_ascii=False, indent=4)
                    basic_logger.info(f'Trying to restore meta.')
                    zipped_meta = self.restore_meta(data_part, path=path)
                    error_count += 1
                    if not zipped_meta[0][0]:
                        error_count +=1
                        basic_logger.critical(f'Critical SHIT happened with {schema} migration. Trying to do ctrl+z....')
                        a = input('Critical_shit. Do ctrl+z?\n')
                        if a == 'y':
                            success = self.ctrl_z(path, params={"filter": f'owner={OWNER}', 'limit': '500'})
                            if success:
                                print('Suppose to check logs.')
                                input('Ctrl+z done, go check logs.')
                            else:
                                input('Critical error, need to check logs.')
                        else:
                            continue
                else:
                    response_json = response.json()
                error_count += 1
            else:
                response_json = response.json()
            if not zipped_meta:
                zipped_meta = zip([x[0] for x in data_part], response_json)
            if schema.lower() != 'salesreturn_from_retail':
                error_count += self.update_db_metas(zipped_meta)


        return error_count
            
    # def make_documents_bonds(self, data, entity, **kwargs):
    #   судя по всему эта хуйня не нужна
    #     path = REQUEST_LINKS.get(entity)
    #     if not path:
    #         return False
    #     self.request('POST', path=path, data=data, **kwargs) #пока что хуйня, смотри выше как формируются данные
        

    def db_migrate(self, **kwargs):
        for entity, path in REQUEST_LINKS.items():
            basic_logger.info(f'Start migrating {entity} entity.')
            errors = self.migrate(schema=entity, path=path, **kwargs)
            if errors == 0:
                basic_logger.info(f'{entity} entity succesefully migrated.')
            else:
                basic_logger.info(f'{entity} was migrated with {errors} errors.')
            input(f'{entity} migrated. Press any key to continue')

    def ctrl_z(self, path, mode='delete', **kwargs):
        """Вот эту вот функцию надо будет переписать как-то покрасивее.
        query и params - взаимоисключающие переменные"""
        if kwargs.get('query') and kwargs.get('params'):
            basic_logger.warning(f'ctrl_z gets mutually exclusive arguments param and query.')
            return False
        if kwargs.get('query'):
            query = self.to_query(kwargs['query'])
        else:
            query = self.to_query({})
        params = {}
        if kwargs.get('params'):
            params = kwargs['params']
        else:
            params = {'filter': query}
        href = path
        errors = 0
        while True:
            data_to_delete = self.request('GET', href, params=params)
            if data_to_delete.status_code == 200:
                try:
                    response_json = data_to_delete.json()
                except requests.JSONDecodeError as e:
                    basic_logger.exception(f'{e}', exc_info=True)
                    return False
                
                metas = [{'meta': x['meta']} for x in response_json.get('rows', [])]
                if metas:
                    response = self.request('POST', '/'.join([path, mode]), data=self.dump_json(metas), params=params)
                    if response.status_code == 200:
                        if response_json['meta'].get('nextHref') and kwargs.get('next_href'):
                            continue
                        else:
                            basic_logger.info("\n".join([x.get("info", 'No info') for x in response.json()]))
                            basic_logger.info(f'Удаление завершено с  {errors} ошибками.')
                            return True
                        
                    elif data_to_delete.status_code == 409:
                        basic_logger.warning('Some object are in use and cannot be deleted.')
                        return False
                    else:
                        basic_logger.info(f'In ctrl_z server give {response.status_code}')
                        errors += 1
                        if errors >= 45:
                            return False
            
                else:
                    basic_logger.warning(f'Ctrl+z dont find any data in MyWH to delete.')
                    basic_logger.debug(f'Request params: {params}')
                    return False

    def restore_zipped_meta(self, schema, **kwargs):
        #Полагаю что время выполнения этого на 9к+ реализациях будет оооооооочень долгим....
        body_schema = self.json_load(f'schemas/{schema.lower()}.json')
        item_type = body_schema["#type"].split('.')[0].split(':')[-1]
        if kwargs.get('data'):
            refs_list = kwargs.get('data')
        else:
            refs_list = [(None, {'description': x.get('#value', {}).get('Ref')}) for x in self.db.get_items(item_type, {"#type": body_schema["#type"], "#mywh": {"$exists": False}})]
        zipped_meta = self.restore_meta(refs_list, **kwargs)
        errors = self.update_db_metas(zipped_meta)
        basic_logger.info(f'Attemp to resore meta in DB end with {errors} errors.')

    def compare_exist_entitys(self, **kwargs):

        for n, (key, value) in enumerate(kwargs.get('data').items()):
            if not key or not value:
                return False
            if value != 'Create':
                response = self.request('GET', '/'.join([kwargs.get('entity'), value]), params=kwargs.get('params') or {})
                if response.status_code == 200:
                    if kwargs.get('search_field'):
                        self.db.update_item(next(self.db.get_items(kwargs.get('collection'), query={'#type': kwargs.get('item_type'), kwargs.get('search_field'): key})).get('#value', {}).get('Ref'), response.json()['meta'])
                    elif kwargs.get('guids'):
                        self.db.update_item(kwargs.get('guids')[n], response.json()['meta'])
                else:
                    basic_logger.warning(f'Something goes wrong. Server get back {response.status_code}')
        

if __name__ =='__main__':
    basic_logger = setup_logger('basic_logger', 'logs/my_wh_api.log', DEBUG)
    json_logger = setup_logger('json_logger', f"logs/{dt.datetime.now().strftime('%Y-%m-%d')}_json.log", DEBUG)
    request_logger = setup_logger('request_logger', f"logs/{dt.datetime.now().strftime('%Y-%m-%d')}_request.log", DEBUG)
    db_logger = setup_logger('db_logger', 'logs/MongoDB_log.log', level=INFO)
    basic_logger.info('Start logging session.')
    database = CrowdGamesDB(db='CrowdGamesActual')
    mywh = MyWHAPI(USERNAME, PSWD, database, headers=HEADERS, max_data_size=MAX_DATA_SIZE, max_item_count=MAX_ITEM_COUNT, timeout_time=TIMEOUT_TIME)
    #response = mywh.get_products(params={'pathName':['Товары интернет-магазинов/Настольные игры/CrowdGames']})
    #mywh.json_to_excel(json.dumps(response), columns=['id', 'name', 'article'])
    a = input('Migrate - press M, go back - B, Restore zipped_meta - R, L - load goods.\n')
    if a =='m':
        mywh.db_migrate(only_posted_documents=True, skip_existing=True, skip_deleted=True)
    if a =='b':
        mywh.ctrl_z('/entity/supply', params={"limit": '500', 'filter':'='.join(['owner', OWNER])})
    if a =='r':
        #data = mywh.json_load('Data/Missed_meta_2022-08-09.json')
        schema = 'MOVE'
        mywh.restore_zipped_meta(schema=schema, path=f'/entity/{schema.lower()}', drop_duplicates=False, next_href=True)
    if a == 'l':
        
        for data_path, entity, item_type in zip(['Data/Compare.json'], ['/entity/product'], ["jcfg:CatalogObject.Номенклатура"]):
            data = mywh.json_load(data_path)
            
            mywh.compare_exist_entitys(data=data, entity=entity, collection='Справочники', item_type=item_type, search_field='#value.Description')
            # for guid, value in data.items():

            #     mywh.db.update_item(guid, value)
    basic_logger.info('Finish')